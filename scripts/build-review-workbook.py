# -*- coding: utf-8 -*-
"""
검수표 엑셀 파일을 만든다.
    npm run review:workbook

먼저 npm run review:packet 을 실행해 docs/review-data.json 이 있어야 한다.
결과: docs/검수표.xlsx
"""
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
DATA = json.loads((ROOT / "docs" / "review-data.json").read_text(encoding="utf-8"))
OUT = ROOT / "docs" / "검수표.xlsx"

FONT = "Arial"
NAVY = "1D4ED8"
HEADER_FILL = PatternFill("solid", fgColor="DBEAFE")
INPUT_FILL = PatternFill("solid", fgColor="FFF9C4")   # 노란색 = 여기에 적어 주세요
TITLE_FONT = Font(name=FONT, size=14, bold=True, color="0F172A")
HEAD_FONT = Font(name=FONT, size=10, bold=True, color=NAVY)
BODY_FONT = Font(name=FONT, size=10)
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def style_header(ws, row, widths):
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.font = HEAD_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = CENTER
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def style_body(ws, first_row, last_row, verdict_cols, height=None):
    """verdict_cols: 노란색 입력란으로 표시할 열 번호들"""
    for row in ws.iter_rows(min_row=first_row, max_row=last_row):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = CENTER if cell.column in verdict_cols else WRAP
            if cell.column in verdict_cols:
                cell.fill = INPUT_FILL
        if height:
            ws.row_dimensions[row[0].row].height = height


def add_verdict_dropdown(ws, col_letter, first_row, last_row):
    dv = DataValidation(
        type="list",
        formula1='"O,△,X"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.prompt = "O = 괜찮음 / △ = 수정 필요 / X = 빼야 함"
    dv.promptTitle = "판정"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


wb = Workbook()

# ─────────────────────────── 1. 사용법 ───────────────────────────
ws = wb.active
ws.title = "사용법"
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 100

rows = [
    ("", ""),
    ("", "리커버핏 Coach 검수표"),
    ("", f"생성일 {DATA['generatedAt']}"),
    ("", ""),
    ("", "■ 이 파일은 무엇인가요"),
    ("", "앱이 사용자에게 내보내는 안전 기준과 운동·영양 내용을 전문가가 확인하기 위한 표입니다."),
    ("", "아래 기준들은 아직 임상 근거로 검증된 값이 아닙니다. 기획 단계에서 정한 값입니다."),
    ("", ""),
    ("", "■ 어떻게 표시하나요"),
    ("", "노란색 칸에만 적으시면 됩니다. 나머지 칸은 그대로 두세요."),
    ("", "'판정' 칸을 누르면 O / △ / X 를 고를 수 있습니다. 직접 입력하셔도 됩니다."),
    ("", ""),
    ("", "        O  →  이대로 괜찮음"),
    ("", "        △  →  수정 필요 (오른쪽 '의견' 칸에 수정안을 적어 주세요)"),
    ("", "        X  →  빼야 함 / 위험함"),
    ("", ""),
    ("", "■ 시트 안내"),
    ("", "  ① 영상후보    유튜브 후보 영상을 보고 앱에 연결할지 고릅니다"),
    ("", "  ② 위험신호    운동을 막는 12개 문항이 적절한지 봅니다"),
    ("", "  ③ 운동목록    운동 86개의 단계 배치와 주의사항을 봅니다"),
    ("", "  ④ 영양        4개 트랙과 탄단지 비율을 봅니다"),
    ("", "  ⑤ 기준값      통증 점수 컷오프 등 숫자 기준을 봅니다"),
    ("", "  ⑥ 진행현황    표시한 개수가 자동으로 세어집니다"),
    ("", ""),
    ("", "■ 다 하신 뒤"),
    ("", "이 파일을 그대로 저장해서 보내 주시면 됩니다."),
]
for r, (a, b) in enumerate(rows, start=1):
    ws.cell(row=r, column=2, value=b).font = BODY_FONT

ws["B2"].font = TITLE_FONT
for key in ("B5", "B9", "B17", "B25"):
    ws[key].font = Font(name=FONT, size=11, bold=True, color=NAVY)
ws["B7"].font = Font(name=FONT, size=10, bold=True, color="B45309")
for key in ("B13", "B14", "B15"):
    ws[key].font = Font(name=FONT, size=10, bold=True)

# ─────────────────────────── 2. 영상후보 ───────────────────────────
ws = wb.create_sheet("영상후보")
ws.append(["운동 id", "운동명", "부위 / 단계", "후보 영상 링크", "채널 · 비고", "판정", "의견"])

CANDIDATES = [
    ("knee-p1-01", "쿼드셋", "무릎 P1", "https://www.youtube.com/watch?v=d5A5LWDOPE0", "Quad Set | Knee Stability Series · 영어"),
    ("knee-p1-01", "쿼드셋", "무릎 P1", "https://www.youtube.com/watch?v=5TUK4uT2nnw", "Physical Therapy Exercises · 영어 · 물리치료사"),
    ("knee-p1-01", "쿼드셋", "무릎 P1", "https://www.youtube.com/watch?v=au62CidApd0", "Ask Doctor Jo · 영어 · 물리치료사"),
    ("knee-p1-01", "쿼드셋", "무릎 P1", "https://www.youtube.com/watch?v=mzTKLYET6QA", "MedBridge · 영어 · 교육기관"),
    ("knee-p1-02", "힐슬라이드", "무릎 P1", "https://www.youtube.com/watch?v=t4HVx5NDeHc", "통증 없이 무릎 굽히는 힐 슬라이드 · 한국어"),
    ("knee-p1-02", "힐슬라이드", "무릎 P1", "https://www.youtube.com/watch?v=m06ilKpj87g", "무릎 전문의 재활운동 · 한국어 · 의사"),
    ("knee-p1-02", "힐슬라이드", "무릎 P1", "https://www.youtube.com/watch?v=W3_BUipgYU4", "Seated Active Heel Slide · 영어"),
    ("knee-p1-03", "누워서 다리 들기 (SLR)", "무릎 P1", "https://www.youtube.com/watch?v=WqI8ln0MpjQ", "[무릎 재활 1주차] · 한국어 · 단계별 시리즈"),
    ("knee-p1-03", "누워서 다리 들기 (SLR)", "무릎 P1", "https://www.youtube.com/watch?v=iwiVVQ22vqw", "누워서 쉽게 따라 할 수 있는 무릎 운동 · 한국어"),
    ("knee-p1-03", "누워서 다리 들기 (SLR)", "무릎 P1", "https://www.youtube.com/watch?v=tNLINRNEQtM", "무릎 통증을 줄여주는 무릎 운동 · 한국어"),
    ("back-p1-02", "골반 틸트", "허리 P1", "https://www.youtube.com/watch?v=SaUhRNOjwUg", "골반 후경사 운동(누운자세) · 한국어 · 앱 자세와 가장 가까움"),
    ("back-p1-02", "골반 틸트", "허리 P1", "https://www.youtube.com/watch?v=HAdd1sP6ruE", "누워서 골반 움직이기 · 한국어"),
    ("back-p1-02", "골반 틸트", "허리 P1", "https://www.youtube.com/watch?v=4j0vN1WEIyg", "Posterior Pelvic Tilt · 영어 · 환자 교육용"),
    ("ankle-p1-02", "발목 펌프", "발목 P1", "https://www.youtube.com/watch?v=BrJ67g0NKqE", "삼성서울병원 스포츠의학실 · 한국어 · 병원 제작"),
    ("ankle-p1-02", "발목 펌프", "발목 P1", "https://www.youtube.com/watch?v=G2Cm-VvaY0w", "[무릎 재활 1주차] 앉아서 발목 펌프 · 한국어"),
    ("ankle-p1-01", "발목 알파벳 · 종아리 스트레칭", "발목 P1", "https://www.youtube.com/watch?v=u5WF34iRvB4", "서울백병원 발목 염좌 재활 · 한국어 · 병원 제작 · 여러 동작 포함"),
    ("ankle-p1-01", "발목 알파벳 · 종아리 스트레칭", "발목 P1", "https://www.youtube.com/watch?v=eJqA2gPdgHQ", "현직 물리치료사 단계별 발목 재활 · 한국어 · 여러 동작 포함"),
    ("ankle-p1-01", "발목 알파벳 · 종아리 스트레칭", "발목 P1", "https://www.youtube.com/watch?v=hE0IsV4e9oE", "발목인대 & 염좌 재활운동 1주차 · 한국어"),
    ("shoulder-p1-01", "펜듈럼 (진자 운동)", "어깨 P1", "https://www.youtube.com/watch?v=aU3hzceroOk", "코드만 운동 정확히 알고 하기 · 한국어"),
    ("shoulder-p1-01", "펜듈럼 (진자 운동)", "어깨 P1", "https://www.youtube.com/watch?v=Pl5XLHKNr30", "시계추운동·진자운동·벽타기 · 한국어"),
    ("shoulder-p1-01", "펜듈럼 (진자 운동)", "어깨 P1", "https://www.youtube.com/watch?v=ICmlLI-rd7g", "물리치료사가 알려주는 어깨 재활 · 한국어 · 물리치료사"),
    ("shoulder-p1-02", "견갑 세팅", "어깨 P1", "https://www.youtube.com/watch?v=T98_YpZwny0", "견갑골 안정화 운동 · 한국어"),
    ("shoulder-p1-02", "견갑 세팅", "어깨 P1", "https://www.youtube.com/watch?v=mH4Ipg1DRak", "어깨건강 예방운동: 견갑골 안정화 · 한국어"),
    ("shoulder-p1-02", "견갑 세팅", "어깨 P1", "https://www.youtube.com/watch?v=LXBNdN5RIIs", "어깨·견갑골 안정화 전거근 운동 · 한국어"),
    ("neck-p1-01", "턱 당기기 (누워서)", "목 P1", "https://www.youtube.com/watch?v=PaIzpXp33Sc", "⚠ 거북목 2분 교정 올바른 친턱 운동법 · 한국어"),
    ("neck-p1-01", "턱 당기기 (누워서)", "목 P1", "https://www.youtube.com/watch?v=bEPVu8K53qQ", "⚠ 일반인 99% 잘못하고 있는 턱 당기기 · 한국어 · 흔한 실수"),
    ("neck-p1-01", "턱 당기기 (누워서)", "목 P1", "https://www.youtube.com/watch?v=scWH5e22WIs", "⚠ 턱 당기기 해야 하나 말아야 하나 · 한국어 · 검수 전 꼭 보세요"),
    ("neck-p1-01", "턱 당기기 (누워서)", "목 P1", "https://www.youtube.com/watch?v=H0r_ntsH4v4", "⚠ 거북목 교정 턱 당기기 제대로 하는법 · 한국어"),
]
for row in CANDIDATES:
    ws.append([*row, "", ""])

last = ws.max_row
style_header(ws, 1, [16, 26, 12, 46, 46, 8, 30])
style_body(ws, 2, last, verdict_cols={6}, height=30)
add_verdict_dropdown(ws, "F", 2, last)
for r in range(2, last + 1):
    ws.cell(row=r, column=4).font = Font(name=FONT, size=9, color="1D4ED8", underline="single")
    ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=7).fill = INPUT_FILL

note = ws.cell(row=last + 2, column=1,
               value="⚠ 자세가 정확한지·안전한지 확인하지 않은 후보입니다. 조회수와 댓글 수는 확인하지 못했습니다. "
                     "⚠ 표시된 '턱 당기기'는 모두에게 맞는 운동이 아니라는 견해가 있어 특히 주의가 필요합니다.")
note.font = Font(name=FONT, size=9, color="B45309")
ws.merge_cells(start_row=last + 2, start_column=1, end_row=last + 2, end_column=7)
ws.row_dimensions[last + 2].height = 30
note.alignment = WRAP

VIDEO_LAST = last

# ─────────────────────────── 3. 위험신호 ───────────────────────────
ws = wb.create_sheet("위험신호")
ws.append(["#", "문항", "보조 설명", "판정", "수정 의견"])
for i, f in enumerate(DATA["redFlags"], start=1):
    ws.append([i, f["label"], f["hint"], "", ""])
last = ws.max_row
style_header(ws, 1, [5, 48, 44, 8, 34])
style_body(ws, 2, last, verdict_cols={4}, height=32)
add_verdict_dropdown(ws, "D", 2, last)
for r in range(2, last + 1):
    ws.cell(row=r, column=5).fill = INPUT_FILL
    ws.cell(row=r, column=1).alignment = CENTER

for offset, text in enumerate([
    "규칙: 위 12개 중 하나라도 해당하면 운동 루틴을 제공하지 않고 전문가 상담 안내로 보냅니다. 점수화하지 않고, 대체 운동도 주지 않습니다.",
    "여쭙고 싶은 것: 빠진 항목이 있습니까? 빼도 되는 항목이 있습니까? 문구가 일반인에게 오해 없이 전달됩니까?",
    "추가로 넣어야 할 항목이 있다면 여기에 적어 주세요 →",
], start=2):
    c = ws.cell(row=last + offset, column=1, value=text)
    c.font = Font(name=FONT, size=9, color="475569")
    ws.merge_cells(start_row=last + offset, start_column=1, end_row=last + offset, end_column=4)
    c.alignment = WRAP
ws.cell(row=last + 4, column=5).fill = INPUT_FILL
ws.cell(row=last + 4, column=5).border = BORDER

REDFLAG_LAST = last

# ─────────────────────────── 4. 운동목록 ───────────────────────────
ws = wb.create_sheet("운동목록")
ws.append(["운동 id", "운동명", "분류", "단계", "난이도", "목적", "방법", "주의사항", "판정", "의견"])

LEVEL = {"beginner": "초급", "intermediate": "중급", "advanced": "상급"}
for e in DATA["exercises"]:
    ws.append([
        e["id"],
        e["name"],
        DATA["categoryLabel"].get(e["category"], e["category"]),
        f"P{e['phase']}" if e["phase"] else "-",
        LEVEL.get(e["level"], e["level"]),
        e["purpose"],
        " → ".join(e["description"]),
        " / ".join(e["precautions"]),
        "",
        "",
    ])
last = ws.max_row
style_header(ws, 1, [17, 24, 16, 6, 7, 40, 60, 44, 8, 28])
style_body(ws, 2, last, verdict_cols={9}, height=44)
add_verdict_dropdown(ws, "I", 2, last)
for r in range(2, last + 1):
    ws.cell(row=r, column=10).fill = INPUT_FILL
    for col in (4, 5):
        ws.cell(row=r, column=col).alignment = CENTER
ws.auto_filter.ref = f"A1:J{last}"

EXERCISE_LAST = last

# ─────────────────────────── 5. 영양 ───────────────────────────
ws = wb.create_sheet("영양")
ws.append(["트랙", "대상", "탄수화물 %", "단백질 %", "지방 %", "핵심 원칙", "학생 안내", "판정", "의견"])
for p in DATA["plans"]:
    m = p.get("macro")
    ws.append([
        p["title"],
        p["targetUser"],
        f"{m['c'][0]}~{m['c'][1]}" if m else "-",
        f"{m['p'][0]}~{m['p'][1]}" if m else "-",
        f"{m['f'][0]}~{m['f'][1]}" if m else "-",
        " / ".join(p["principles"]),
        " / ".join(p["studentGuide"]),
        "",
        "",
    ])
last = ws.max_row
style_header(ws, 1, [22, 34, 11, 10, 9, 56, 50, 8, 28])
style_body(ws, 2, last, verdict_cols={8}, height=70)
add_verdict_dropdown(ws, "H", 2, last)
for r in range(2, last + 1):
    ws.cell(row=r, column=9).fill = INPUT_FILL
    for col in (3, 4, 5):
        ws.cell(row=r, column=col).alignment = CENTER

c = ws.cell(row=last + 2, column=1,
            value="여쭙고 싶은 것: 탄단지 비율이 최신 한국인 영양소 섭취기준과 맞습니까? "
                  "성장기 청소년에게 그대로 적용해도 됩니까? 빠진 주의사항이 있습니까?")
c.font = Font(name=FONT, size=9, color="475569")
ws.merge_cells(start_row=last + 2, start_column=1, end_row=last + 2, end_column=9)
c.alignment = WRAP

NUTRITION_LAST = last

# ─────────────────────────── 6. 기준값 ───────────────────────────
ws = wb.create_sheet("기준값")
ws.append(["구분", "기준", "앱의 반응", "판정", "의견"])
t = DATA["thresholds"]
CRITERIA = [
    ("설문 시점", f"통증 {t['painStop']}점 이상", "운동 제공 안 함, 전문가 상담 안내"),
    ("설문 시점", f"통증 {t['painPhase1']}~{int(t['painStop'])-1}점", "Phase 1 (통증 완화 / 가동성 회복)"),
    ("설문 시점", f"통증 {t['painPhase2']}~{int(t['painPhase1'])-1}점", "Phase 2 (안정화 / 근육 활성화)"),
    ("설문 시점", "통증 0점", "회복운동 대상 아님. 목표별 루틴"),
    ("운동 중", f"통증 0~{int(t['inWorkoutReduce'])-1}점", "현재 루틴 유지"),
    ("운동 중", f"통증 {t['inWorkoutReduce']}~{int(t['inWorkoutStop'])-1}점", "횟수 절반, 가동범위 축소, 쉬운 동작으로"),
    ("운동 중", f"통증 {t['inWorkoutStop']}점 이상", "해당 운동 중단, 기록 저장, 상담 안내"),
    ("운동 후", "통증이 운동 전보다 높음", "다음 루틴 강도 하향"),
    ("위험도", "위험 신호 1개 이상 또는 통증 7점 이상", "Red — 운동 제공 중단"),
    ("위험도", "통증 4~6점 또는 붓기·당일 발생·상시 통증", "Yellow — 저강도 회복운동"),
    ("위험도", "통증 0~3점, 위험 신호 없음", "Green — 회복·교정 운동"),
    ("위험도", "통증 0점 + 주 3회 이상 운동 + 수행능력 목표", "Performance — 기능성 운동"),
]
for row in CRITERIA:
    ws.append([*row, "", ""])
last = ws.max_row
style_header(ws, 1, [12, 40, 46, 8, 34])
style_body(ws, 2, last, verdict_cols={4}, height=26)
add_verdict_dropdown(ws, "D", 2, last)
for r in range(2, last + 1):
    ws.cell(row=r, column=5).fill = INPUT_FILL

c = ws.cell(row=last + 2, column=1,
            value="여쭙고 싶은 것: 차단 기준 7점이 타당합니까? 더 낮춰야 합니까? "
                  "3개월 이상 지속되는 통증은 점수와 무관하게 차단해야 합니까? (현재는 점수만 봅니다)")
c.font = Font(name=FONT, size=9, color="475569")
ws.merge_cells(start_row=last + 2, start_column=1, end_row=last + 2, end_column=5)
c.alignment = WRAP

CRITERIA_LAST = last

# ─────────────────────────── 7. 진행현황 ───────────────────────────
ws = wb.create_sheet("진행현황")
ws.append(["시트", "전체", "O", "△", "X", "남음"])
sheets = [
    ("영상후보", "F", 2, VIDEO_LAST),
    ("위험신호", "D", 2, REDFLAG_LAST),
    ("운동목록", "I", 2, EXERCISE_LAST),
    ("영양", "H", 2, NUTRITION_LAST),
    ("기준값", "D", 2, CRITERIA_LAST),
]
for r, (name, col, first, lastrow) in enumerate(sheets, start=2):
    total = lastrow - first + 1
    rng = f"{name}!${col}${first}:${col}${lastrow}"
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=total)
    ws.cell(row=r, column=3, value=f'=COUNTIF({rng},"O")')
    ws.cell(row=r, column=4, value=f'=COUNTIF({rng},"△")')
    ws.cell(row=r, column=5, value=f'=COUNTIF({rng},"X")')
    ws.cell(row=r, column=6, value=f"=B{r}-C{r}-D{r}-E{r}")

total_row = len(sheets) + 2
ws.cell(row=total_row, column=1, value="합계")
for col in range(2, 7):
    letter = get_column_letter(col)
    ws.cell(row=total_row, column=col, value=f"=SUM({letter}2:{letter}{total_row-1})")

style_header(ws, 1, [14, 8, 8, 8, 8, 8])
style_body(ws, 2, total_row, verdict_cols=set(range(2, 7)), height=20)
for col in range(1, 7):
    ws.cell(row=total_row, column=col).font = Font(name=FONT, size=10, bold=True)
    ws.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="E2E8F0")

c = ws.cell(row=total_row + 2, column=1, value="표시하실 때마다 자동으로 세어집니다. (엑셀에서 열면 갱신됩니다)")
c.font = Font(name=FONT, size=9, color="475569")

wb.save(OUT)
print(f"저장 완료: {OUT}")
print(f"  영상후보 {VIDEO_LAST-1}행 · 위험신호 {REDFLAG_LAST-1}행 · 운동 {EXERCISE_LAST-1}행 · 영양 {NUTRITION_LAST-1}행 · 기준값 {CRITERIA_LAST-1}행")
